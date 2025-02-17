import os
import re
import zipfile
import tempfile
import shutil
import hashlib
from pynput import keyboard
import xml.etree.ElementTree as ET
from PIL import Image
from tqdm import tqdm
from openpyxl import Workbook  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

import AutoUpdate
latest_version = ""
url = "https://api.github.com/repos/KRNtr/batch-analyser/contents/releases"
download_url = f"https://raw.githubusercontent.com/KRNtr/batch-analyser/main/releases/batch_analyser_{latest_version}.exe"

os.system('color')

class COLOUR:
    ESC = '\x1b'
    GREEN  = ESC + '[32m'
    GREEN_BG  = ESC + '[42m'
    RED    = ESC + '[31m'
    RED_BG    = ESC + '[41m'
    YELLOW = ESC + '[33m'
    CYAN = "\033[0;36m"
    STOP   = '\x1b[0m'

def write_errors_to_xlsx(all_errors, path, batch_name, filename='Validation_Errors'):
    if not os.path.exists(path):
        os.makedirs(path)
    output_filename = f"{filename}_{batch_name}.xlsx"
    output_file_path = os.path.join(path, output_filename)
    wb = Workbook()
    ws = wb.active
    ws.append(['UPC', 'Error'])  # Updated header for clarity
    for error_details in all_errors:
        upc = error_details['folder_name']
        errors = error_details['error']
        if isinstance(errors, list):
            for error in errors:
                ws.append([upc, error])
        else:
            ws.append([upc, errors])
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number (e.g., A=1, B=2)
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust the width factor as necessary
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    wb.save(output_file_path) 
    return output_file_path

def check_for_duplicates(file_list):
    my_list = set()
    duplicates = False
    for line in file_list:
        clean_line = line.strip()
        if clean_line in my_list:
            duplicates = True
            break
        my_list.add(clean_line)

    return duplicates

def check_for_errors(zip_file_path):

    global stop_execution

    # Create a temporary directory to extract the zip file
    temp_dir = tempfile.mkdtemp(prefix = "batch_analyser_temp_")
    failed_subdirs = []
    error_details = []
    xml_name = ""
    xml_exists = True

    try:
        # Extract all contents of the zip file into the temporary directory
        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        if stop_execution:
            return
        extracted_items = os.listdir(temp_dir)
        batch_id = extracted_items[0]
        if len(extracted_items) != 1 or not os.path.isdir(os.path.join(temp_dir, extracted_items[0])):
            print(COLOUR.RED + "Error: The .zip file should contain exactly one parent folder" + COLOUR.STOP)
            return
        parent_folder_path = os.path.join(temp_dir, extracted_items[0])
        batch_name = os.path.basename(zip_file_path)
        batch_name = batch_name.replace(".zip", "")

        match = re.match(r'^(.*?)_', batch_name)
        if match:
            distributor_name = match.group(1)
            print(f"Distributor name: {distributor_name}")
        else:
            print(COLOUR.RED + "Distributor name not found." + COLOUR.STOP)
        
        if distributor_name == "Netease":
            match = re.search(r'_(\w+?)_', batch_name)
            if match:
                batch_type = match.group(1)  # Extracts the batch type before the second underscore
                print("Batch Type: " + batch_type)
            else:
                print(COLOUR.RED + "Error: Batch type not found." + COLOUR.STOP)
        elif distributor_name == "Tuned":
            match = re.search(r"_([A-Za-z]+)_\d+\.zip", batch_name)
            if match:
                batch_type = match.group(1)  # "Insert"
                print("Batch Type:", batch_type)
            else:
                print(COLOUR.RED + "Error: Batch type not found." + COLOUR.STOP)
    
        else:
            match = re.search(r'_(\w+)_', batch_name)
            if match:
                batch_type = match.group(1) 
                print("Batch Type:", batch_type)
            else:
                print(COLOUR.RED + "Error: Batch type not found." + COLOUR.STOP)
        
        if stop_execution:
            return

        # Get total number of subdirectories (for progress tracking)
        subdirs = [subdir_name for subdir_name in os.listdir(parent_folder_path) if os.path.isdir(os.path.join(parent_folder_path, subdir_name))]
        resource_folder = True
        
        # Progress bar: Wrap the iteration with tqdm to show progress
        for subdir_name in tqdm(subdirs, desc="Analyzing batch files... ", unit=" folder", bar_format="{l_bar}{bar:30}{r_bar}", ncols=150):
            if stop_execution:
                return
            subdir_path = os.path.join(parent_folder_path, subdir_name)
            res_dir = os.path.join(subdir_path, "resources")
            errors = []
            upc_txt_list = []
            filename_txt_list = []

            if os.path.isdir(subdir_path):
                # Check for various errors
                xml_files = [f for f in os.listdir(subdir_path) if f.endswith('.xml')]
                xml_name = xml_files[0]
                if not xml_files:
                    xml_exists = False
                    errors.append("Missing metadata .xml file")
                
                if not os.path.exists(os.path.join(subdir_path, "resources")):
                    resource_folder = False
                else:
                    resource_folder = True

                if batch_type == "Insert":

                    upc_txt_filename = f"{subdir_name}.txt"
                    upc_txt_path = os.path.join(subdir_path, upc_txt_filename)
                    filenamelist_path = os.path.join(subdir_path, "filenamelist.txt")

                    # Count lines in 'upc'.txt and filename.txt, then identify mismatches
                    try:
                        with open(upc_txt_path, 'r', encoding='utf-8') as f_upc:
                            lines = f_upc.readlines() 
                            for line in lines:
                                if line != '\n':
                                    upc_txt_list.append(line)
                            num_upc_lines = len(upc_txt_list)
                    except FileNotFoundError:
                        errors.append(f"Missing {upc_txt_filename}")

                    try:
                        with open(filenamelist_path, 'r', encoding='utf-8') as f_list:
                            lines = f_list.readlines() 
                            for line in lines:
                                if line != '\n':
                                    filename_txt_list.append(line)
                            num_filename_lines = len(filename_txt_list)
                    except FileNotFoundError:
                        errors.append("Missing filenamelist.txt")
                    
                    if num_upc_lines != num_filename_lines:
                        errors.append("Line count mismatch")

                    # Check for any duplicate entries in the upc txt and filename txt files
                    if check_for_duplicates(upc_txt_list):
                        errors.append("Duplicate lines upc.txt file")
                    if check_for_duplicates(filename_txt_list):
                        errors.append("Duplicate lines filenamelist.txt file")
                
                if errors:
                    failed_subdirs.append(subdir_name)
                    error_details.append({'folder_name': subdir_name, 'error': ', '.join(errors)})

        if stop_execution:
            return    
        
        print("==== release file integrity checked ====")

        if batch_type != "Takedown":                
            for subdir_name in tqdm(subdirs, desc="Checking for corrupted JPEGs... ", unit=" folder", bar_format="{l_bar}{bar:30}{r_bar}", ncols=150):
                if stop_execution:
                    return
                subdir_path = os.path.join(parent_folder_path, subdir_name)
                res_dir = os.path.join(subdir_path, "resources")
                
                if resource_folder:
                    img_path = os.path.join(res_dir, (subdir_name + ".jpg"))
                else:
                    if distributor_name == "AWA":
                        img_path = os.path.join(subdir_path, subdir_name + "_cover.jpg")
                    else:
                        img_path = os.path.join(subdir_path, subdir_name + ".jpg")
                
                jpg_errors = []
                image_load = False
                try:
                    # Verify integrity of the image
                    with Image.open(img_path) as art:
                        art.verify()
                    image_load = True
                except (IOError, SyntaxError):
                    jpg_errors.append("Error loading album art")
                
                if xml_exists and image_load:
                    hasher = hashlib.md5()
                    with open(img_path, "rb") as open_image:
                        content = open_image.read()
                        hasher.update(content)
                    img_hashsum = hasher.hexdigest()

                    if xml_name == "metadata.xml":
                        try:
                            # Try to parse metadata.xml (or another XML file)
                            tree = ET.parse(os.path.join(subdir_path, xml_name))
                            root = tree.getroot()

                            # Extract namespace dynamically
                            namespace = {}
                            for elem in root.iter():
                                if '}' in elem.tag:
                                    namespace_uri = elem.tag.split('}')[0].strip('{')
                                    namespace[namespace_uri] = namespace_uri

                            if namespace:
                                # Now, use the extracted namespace to search for the checksum
                                namespace_uri = list(namespace.values())[0]  # Use the first found namespace

                                # Try to find the checksum in the metadata.xml structure
                                image_details = root.find(f".//{{{namespace_uri}}}album/{{{namespace_uri}}}artwork_files/{{{namespace_uri}}}file/{{{namespace_uri}}}checksum")
                                if image_details is not None:
                                    xml_hashsum = image_details.text
                                    if img_hashsum != xml_hashsum:
                                        jpg_errors.append(f"Album art HashSum mismatch in metadata.xml")
                                else:
                                    jpg_errors.append("Album art HashSum not found in metadata.xml")
                                
                        except FileNotFoundError:
                            jpg_errors.append("metadata.xml not found")
                        except ET.ParseError:
                            jpg_errors.append("Error parsing metadata.xml")

                    else:
                        # Now check other XML files (like {subdir_name}.xml) with the second approach
                        try:
                            tree = ET.parse(os.path.join(subdir_path, subdir_name+'.xml'))
                            root = tree.getroot()
                            image_details = None
                            image_details = root.find(".//ImageDetailsByTerritory/TechnicalImageDetails/File/HashSum/HashSum")
                            if image_details is not None:
                                xml_hashsum = image_details.text
                                if img_hashsum != xml_hashsum:
                                    jpg_errors.append("Album art HashSum mismatch")

                        except FileNotFoundError:
                            jpg_errors.append(f"{subdir_name}.xml not found")
                        except ET.ParseError:
                            jpg_errors.append(f"Error parsing {subdir_name}.xml")

                else:
                    jpg_errors.append("Missing Album Art")

                if len(jpg_errors) > 0:
                    failed_subdirs.append(subdir_name)
                    error_details.append({'folder_name': subdir_name, 'error': jpg_errors})

        if stop_execution:
            return

            print("==== release art integrity checked ====\n")

        if stop_execution:
            return
        elif error_details:

            # Create directories for error folders
            failed_releases_folder = os.path.join(os.path.dirname(zip_file_path), "Failed_Releases_" + batch_name)
            os.makedirs(failed_releases_folder, exist_ok=True)

            # Print errors and move failed subdirectories to their respective error folders
            for error in error_details:
                print(COLOUR.YELLOW + f"{error['folder_name']} - Issue: {error['error']}" + COLOUR.STOP)
                shutil.move(os.path.join(parent_folder_path, error['folder_name']), failed_releases_folder)

            print("\n---------------------------------------\n")

            # Remove the failed upc's from the batch by creating a new zipped directory and giving the option to delete the original
            while True:
                if stop_execution:
                    break
                create_cleaned = input("Would you like to create a new batch folder cleared of failed releases? " + COLOUR.YELLOW + "(y/n)" + COLOUR.STOP + ": ").lower()

                if create_cleaned == 'y':
                    print(" ")
                    ask_delete = True
                    while True:
                        new_zip_batch_name = input("Enter a name for the new batch (press " + COLOUR.YELLOW + "enter" + COLOUR.STOP + " to leave name unchanged or " + COLOUR.YELLOW + "+" + COLOUR.STOP + " or " + COLOUR.YELLOW + "-" + COLOUR.STOP + " to add/subtract batch id by 1): ")
                        if new_zip_batch_name == "+" or "-" or len(new_zip_batch_name) == 0:
                            break

                    # Get current batch ID from the filename
                    if distributor_name == "Netease":
                        match = re.search(r'_(\d{8}_\d{6})(?=\.zip)', batch_name)
                        if match:
                            old_zip_id = match.group(1)
                    else:
                        match = re.search(rf'{batch_type}_(\d+)\.zip$', os.path.basename(zip_file_path))
                        if match:
                            old_zip_id = match.group(1)

                    if len(new_zip_batch_name) == 0:
                        new_zip_batch_name = os.path.basename(zip_file_path)
                        ask_delete = False
                    
                    elif new_zip_batch_name == "+" or "-":
                        zip_id_last_one = int(old_zip_id[-1])  # Get last digit as integer

                        if batch_id.isdigit():
                            batch_id_last_one = int(batch_id[-1])
                        else:
                            match = re.search(r'(\d+)$', batch_id)
                            if match:
                                # Extract the last digits
                                batch_id_last_one = int((match.group(1))[-1])
                                
                        # Increment the last digit, or reset to 0 if it's already 9
                        if new_zip_batch_name == "+":                            
                            if zip_id_last_one < 9:
                                zip_id_last_one += 1
                            else:
                                zip_id_last_one = 0
                            
                            if batch_id_last_one < 9:
                                batch_id_last_one += 1
                            else:
                                batch_id_last_one = 0
                        elif new_zip_batch_name == "-":
                            if zip_id_last_one > 0:
                                zip_id_last_one -= 1
                            else:
                                zip_id_last_one = 9
                            
                            if batch_id_last_one > 0:
                                batch_id_last_one -= 1
                            else:
                                batch_id_last_one = 9
                                    
                        new_zip_id = old_zip_id[:-1] + str(zip_id_last_one)
                        new_batch_id = batch_id[:-1] + str(batch_id_last_one)
                        new_zip_batch_name = batch_name.replace(old_zip_id, new_zip_id)
                        new_parent_folder = batch_id.replace(batch_id, new_batch_id)
                        cleaned_zip_path = zip_file_path.replace(old_zip_id, new_zip_id)

                    try:
                        new_parent_folder_path = parent_folder_path.replace(batch_id, new_batch_id)
                        os.rename(parent_folder_path, new_parent_folder_path)
                    except Exception as e:  # Use a general exception to catch any issues
                        print(f"Error: {str(e)}")

                    try:
                        # Verify extracted batch complete path and existence
                        batch_complete_name = [file for file in os.listdir(new_parent_folder_path) if file.startswith("BatchComplete") or file.endswith(".complete")]
                        
                        if batch_complete_name:
                            batch_complete_path = os.path.join(new_parent_folder_path, batch_complete_name[0])
                            
                            if batch_complete_name[0] != "delivery.complete":
                                batch_complete = batch_complete_name[0].replace(batch_id, new_batch_id)
                            else:
                                batch_complete = batch_complete_name[0]
                            
                            new_batch_complete_path = os.path.join(new_parent_folder_path, batch_complete)
                            if os.path.exists(batch_complete_path ) and batch_complete != "delivery.complete":
                                os.rename(batch_complete_path, new_batch_complete_path)

                    except FileNotFoundError as fnf_error:
                        print(f"File not found error: {fnf_error}")
                    except Exception as e:
                        print(f"Unexpected error during rename: {e}")

                    # Update parent_folder_path
                    files_to_zip = []
                    for subdir_name in os.listdir(new_parent_folder_path):
                        if subdir_name not in failed_subdirs:
                            subdir_path = os.path.join(new_parent_folder_path, subdir_name)
                            for root, _, files in os.walk(subdir_path):
                                for file in files:
                                    full_path = os.path.join(root, file)
                                    files_to_zip.append(full_path)
                    
                    #new_batch_complete_path = os.path.join(new_parent_folder_path, re)
                    files_to_zip.append(new_batch_complete_path)

                    with zipfile.ZipFile(cleaned_zip_path, 'w', zipfile.ZIP_DEFLATED) as cleaned_zip:
                        for file in tqdm(files_to_zip, desc="Creating cleaned zip", unit="file", bar_format="{l_bar}{bar:30}{r_bar}", ncols=100):
                            arcname = os.path.relpath(file, temp_dir)
                            cleaned_zip.write(file, arcname)

                    print("\n" + COLOUR.YELLOW + f"Cleaned zip file created: {cleaned_zip_path}\n" + COLOUR.STOP)

                    # Ask if the user wants to delete the original zip file
                    if ask_delete == True:
                        delete_original = input("Would you like to delete the original zip file? " + COLOUR.YELLOW + "(y/n)" + COLOUR.STOP + ": ").lower()
                        if delete_original == keyboard.Key("esc"):
                            stop_execution = True
                            print("\nStopping action...")
                        elif delete_original == 'y':
                            os.remove(zip_file_path)
                            print(COLOUR.YELLOW + f"Original zip file deleted: {zip_file_path}\n" + COLOUR.STOP)
                elif create_cleaned == 'n':
                    print("\nNo cleaned zip file created.")
                    break
                else:
                    print("Invalid input. Please enter 'y' or 'n'.")
            
            if not stop_execution:
                validation_results_path = write_errors_to_xlsx(error_details, failed_releases_folder, batch_name)
                if os.path.exists(validation_results_path):
                    print(COLOUR.GREEN + "Validation results written to xlsx file" + COLOUR.STOP)
                elif stop_execution:
                    return
                else:
                    print(COLOUR.RED + "Error: Validation results not written to xlsx file" + COLOUR.STOP)

        # If batch is good, show positive message
        else:
            print(COLOUR.GREEN + "All releases are valid. No issues found." + COLOUR.STOP)        
    
    except stop_execution:
        print("\n Operation ended by user.")

    # Clean up temporary directory
    finally:
        shutil.rmtree(temp_dir)
        print("\nTemp files cleared.")
        return error_details

stop_execution = False

def on_press(key):
    """Set the global stop_execution flag if the Escape key is pressed."""
    global stop_execution
    if key == keyboard.Key.esc:
        stop_execution = True
        print("\nEscape key detected. Exiting...")

def start_key_listener():
    """Start a non-blocking key listener."""
    listener = keyboard.Listener(on_press=on_press)
    listener.start()
    return listener

def main():
    os.system('cls')
    global latest_version
    exe_name = os.path.basename(__file__)
    local_dir = os.path.dirname(__file__)
    local_version = AutoUpdate.get_local_version(exe_name)

    if local_version == "":
        local_version = "v.undefined"
        print(f"\n{COLOUR.RED}Local version could not be determined - could not check for updates.\nCurrent version may be outdated. Please contact administrator.\n")
    else:
        latest_version = AutoUpdate.get_latest_version(url)
        if latest_version is not None:
            AutoUpdate.check_for_updates(local_version, local_dir, latest_version, download_url)
        else:
            print(f"\n{COLOUR.RED}Could not check for updates, please restart tool.\nIf issue persists contact Jake or Kai.\n")
    
    print("\n\n▐░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░▌")
    print("▐░░░░░░░░░ batch_analyser_" + local_version +" ░░░░░░░░░▌")
    print("▐░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░░▌")

    print(f"{COLOUR.GREEN}This is version v1.3.5{COLOUR.STOP}")

    print("\nThis tool checks a .zip batch file for issues that may cause errors when processed by Sendsei.\n")
    print("It will check:\n"
        "    - that the contents of each release folder is correct\n"
        "    - that all the required files / folders are present\n"
        "    - that the download files contain the right information\n"
        "    - the integrity of album art\n")
    print("When the check is over, the tool will:\n"
        "    - move the failed releases into a folder named 'batch_name'_Failed Releases (if any are detected)\n"
        "    - create a xlsx file containing issue information\n"
        "    - give you the option to create a new batch cleared of the failed releases\n"
        "    - give you the option to delete the original batch\n")
    
    print(COLOUR.YELLOW + "Press ESC at any time to safely exit the tool\n  - The tool makes temporary files to function\n  - Exiting with the esc key ensures these files are deleted\n" + COLOUR.STOP)
    global stop_execution
    listener = start_key_listener()

    try:
        while not stop_execution:
            zip_file_path = input(COLOUR.CYAN + "+++ input batch.zip file path here or press " + COLOUR.YELLOW + "(esc)" + COLOUR.CYAN + " to exit:   " + COLOUR.STOP).strip()
            if len(zip_file_path) != 0:
                zip_file_path = (zip_file_path.replace('"', "").replace("'", "").replace("& ", "").strip())
                while not os.path.exists(zip_file_path):
                    print(COLOUR.RED + f"Error: The file '{zip_file_path}' does not exist." + COLOUR.STOP)
                    zip_file_path = input("Please enter a valid zip file path: ").strip()

                print("\n---------------------------------------\n")
                check_for_errors(zip_file_path)
                print("\n---------------------------------------\n")

    finally:
        listener.stop()
        print("Exiting...\n\n")
    
    input("Press enter to exit: ")

if __name__ == "__main__":
    main()

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~      TO DO         ~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#
#   >>> Add check to see which errors are present before making failed releases sub-directory.
#
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~   KNOWN ISSUES     ~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
#
#   >>> JPEG integrity check can let invalid files through if batch does not get checksum - NOT PERFECT
#           - Should help to find issues that cause batches getting stuck / not being sent in SendSei due to corrupted artwork
#           - Cannot detect artwork that opens but is visually corrupted if no checksum is present
#
#           Version History
#
# v0.9
# 21/10/2024
# Current version will seek issues with release folders in a zipped batch folder.
#   >>> Ensures all files are present
#   >>> Ensures the contents of download files is correct
#   >>> Moves failed releases to new directory making it easy to understand errors
# Progress bars to show that the tool is working, can take a while
# Not publicly tested

# v1.0
# 29/10/2024
# Script will now check for duplicated lines in the 'upc.txt' and 'filename.txt' files in each release folder.
# Adjusted line count functionality to remove whitespace characters, ensuring that it only counts each entry in the files

# v1.1
# 04/11/2024
# Added functionality to scan artwork and check integrity
#   >>> Checks if file will load
#   >>> If batch needs checksum - Checks JPEG hashsum against hashsum specified in metadata xml
# Fixed file scanning to include batches that have no resource folder and instead have artwork in main folder
# Turned script into a loop to improve user experience / efficiency scanning multiple batches

# v1.3.2
# 11/11/2024
# Added functionality for AWA and Tuned
# Improved wording of the CLI to improve user experience
# Added ability to input "-" to decrement the name of new zip file.

# v1.3.4
# 19/11/2024
# Stopped creation of error folders
#   >>> Moves all failed releases to "Failed Releases" folder
#   >>> Writes errors to xlsx file to provide extra infomration
# Added ability to safely stop the program at any time by pressing "esc"
#   >>> Program will still delete temp files if exited this way